import unittest
from pathlib import Path

from helpers import temporary_directory
from shift_positions import ShiftPositionsError, apply_shift, apply_spreadsheet_insert, plan_shift, plan_spreadsheet_insert


class ShiftPositionsTests(unittest.TestCase):
	def test_plans_positive_shift_from_selected_file_onward(self):
		with temporary_directory() as temp_dir:
			folder = Path(temp_dir)
			_create_files(folder, "0001 - A.cbz", "0004 - C.cbz", "0005 - D.cbz")

			plan = plan_shift(folder, "0004 - C.cbz")

			self.assertEqual(
				[
					("0005 - D.cbz", "0006 - D.cbz"),
					("0004 - C.cbz", "0005 - C.cbz"),
				],
				[(item.source_path.name, item.destination_path.name) for item in plan],
			)

	def test_applies_positive_shift_without_overwriting(self):
		with temporary_directory() as temp_dir:
			folder = Path(temp_dir)
			_create_files(folder, "0004 - C.cbz", "0005 - D.cbz")

			apply_shift(plan_shift(folder, "0004 - C.cbz"))

			self.assertFalse((folder / "0004 - C.cbz").exists())
			self.assertTrue((folder / "0005 - C.cbz").exists())
			self.assertTrue((folder / "0006 - D.cbz").exists())

	def test_rejects_target_collision_outside_shifted_range(self):
		with temporary_directory() as temp_dir:
			folder = Path(temp_dir)
			_create_files(folder, "0003 - Existing.cbz", "0004 - C.cbz")

			with self.assertRaises(ShiftPositionsError):
				plan_shift(folder, "0004 - C.cbz", increment=-1)

	def test_rejects_zero_increment(self):
		with temporary_directory() as temp_dir:
			folder = Path(temp_dir)
			_create_files(folder, "0004 - C.cbz")

			with self.assertRaises(ShiftPositionsError):
				plan_shift(folder, "0004 - C.cbz", increment=0)

	def test_plans_spreadsheet_insert_at_position_plus_header(self):
		with temporary_directory() as temp_dir:
			spreadsheet_path = Path(temp_dir) / "order.xlsx"
			_create_workbook(spreadsheet_path)

			plan = plan_spreadsheet_insert(spreadsheet_path, "Issue Release Order", start_position=4, row_count=2)

			self.assertEqual(5, plan.insert_before_row)
			self.assertEqual(2, plan.row_count)
			self.assertEqual(spreadsheet_path.with_name("order.backup.xlsx"), plan.backup_path)

	def test_applies_spreadsheet_insert_and_creates_backup(self):
		with temporary_directory() as temp_dir:
			spreadsheet_path = Path(temp_dir) / "order.xlsx"
			_create_workbook(spreadsheet_path)

			apply_spreadsheet_insert(
				plan_spreadsheet_insert(spreadsheet_path, "Issue Release Order", start_position=2, row_count=1)
			)

			from openpyxl import load_workbook

			workbook = load_workbook(spreadsheet_path)
			try:
				sheet = workbook["Issue Release Order"]
				self.assertEqual("Run", sheet.cell(row=1, column=1).value)
				self.assertEqual("Run 1", sheet.cell(row=2, column=1).value)
				self.assertIsNone(sheet.cell(row=3, column=1).value)
				self.assertEqual("Run 2", sheet.cell(row=4, column=1).value)
			finally:
				workbook.close()

			self.assertTrue((Path(temp_dir) / "order.backup.xlsx").exists())


def _create_files(folder: Path, *names: str) -> None:
	for name in names:
		(folder / name).write_bytes(b"comic")


def _create_workbook(path: Path) -> None:
	from openpyxl import Workbook

	workbook = Workbook()
	sheet = workbook.active
	sheet.title = "Issue Release Order"
	sheet.append(["Run", "Volume", "Issue"])
	sheet.append(["Run 1", 1, "#1"])
	sheet.append(["Run 2", 1, "#2"])
	sheet.append(["Run 3", 1, "#3"])
	workbook.save(path)
	workbook.close()


if __name__ == "__main__":
	unittest.main()
