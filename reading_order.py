"""Read the configured spreadsheet's issue-release order sheet."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from issue_numbers import comparable_issue_number, normalize_issue_label, normalize_volume_label
from models import ReadingOrderEntry


class ReadingOrderError(ValueError):
	pass


def read_reading_order(
	spreadsheet_path: str | Path,
	sheet_name: str,
	issue_overrides: dict[str, dict[str, str]] | None = None,
) -> tuple[ReadingOrderEntry, ...]:
	try:
		from openpyxl import load_workbook
	except ImportError as exc:
		raise ReadingOrderError("openpyxl is required to read .xlsx files. Install dependencies from requirements.txt.") from exc

	workbook = load_workbook(spreadsheet_path, data_only=True, read_only=True)
	try:
		if sheet_name not in workbook.sheetnames:
			raise ReadingOrderError(f"Spreadsheet does not contain sheet '{sheet_name}'")

		return read_entries_from_sheet(workbook[sheet_name], issue_overrides or {})
	finally:
		workbook.close()


def read_entries_from_sheet(sheet: Any, issue_overrides: dict[str, dict[str, str]] | None = None) -> tuple[ReadingOrderEntry, ...]:
	overrides = issue_overrides or {}
	rows = sheet.iter_rows(values_only=True)

	try:
		header_row = next(rows)
	except StopIteration as exc:
		raise ReadingOrderError("Reading-order sheet is empty") from exc

	headers = _header_indexes(header_row)
	if "run" not in headers or "issue" not in headers or "volume" not in headers:
		raise ReadingOrderError("Reading-order sheet must contain 'Run', 'Volume', and 'Issue' columns")

	entries: list[ReadingOrderEntry] = []
	for sheet_row_number, row in enumerate(rows, start=2):
		run = _cell(row, headers["run"])
		raw_volume = _cell(row, headers["volume"])
		raw_issue = _cell(row, headers["issue"])
		if run is None and raw_volume is None and raw_issue is None:
			continue
		if run is None or raw_volume is None or raw_issue is None:
			raise ReadingOrderError(f"Reading-order row {sheet_row_number} must contain Run, Volume, and Issue")

		run_name = str(run).strip()
		volume = normalize_volume_label(raw_volume)
		sequence_number = comparable_issue_number(raw_issue)
		issue_label = overrides.get(run_name, {}).get(sequence_number, normalize_issue_label(raw_issue))
		entries.append(ReadingOrderEntry(position=sheet_row_number - 1, run=run_name, volume=volume, issue_label=issue_label))

	return tuple(entries)


def _header_indexes(header_row: tuple[Any, ...]) -> dict[str, int]:
	headers: dict[str, int] = {}
	for index, header in enumerate(header_row):
		if header is None:
			continue
		headers[str(header).strip().casefold()] = index

	return headers


def _cell(row: tuple[Any, ...], index: int) -> Any:
	if index >= len(row):
		return None

	return row[index]
