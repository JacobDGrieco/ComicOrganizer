"""Shift destination filename position prefixes from a selected file onward."""

from __future__ import annotations

import argparse
import re
import shutil
import sys
from dataclasses import dataclass
from pathlib import Path

from config import ConfigError, load_config


POSITION_PREFIX_RE = re.compile(r"^(?P<position>\d{4}) - (?P<rest>.+)$")


class ShiftPositionsError(ValueError):
	pass


@dataclass(frozen=True)
class ShiftPlanItem:
	source_path: Path
	destination_path: Path
	source_position: int
	destination_position: int


@dataclass(frozen=True)
class SpreadsheetInsertPlan:
	spreadsheet_path: Path
	sheet_name: str
	insert_before_row: int
	row_count: int
	backup_path: Path


def plan_shift(folder: str | Path, from_file: str | Path, increment: int = 1) -> tuple[ShiftPlanItem, ...]:
	"""Build and validate a rename plan without changing files."""
	if increment == 0:
		raise ShiftPositionsError("Increment must not be 0")

	folder_path = Path(folder)
	if not folder_path.is_dir():
		raise ShiftPositionsError(f"Folder does not exist: {folder_path}")

	from_path = _resolve_from_file(folder_path, from_file)
	start_position = _position_from_name(from_path.name)
	if start_position is None:
		raise ShiftPositionsError(f"From file does not start with a 4-digit position prefix: {from_path.name}")
	if not from_path.is_file():
		raise ShiftPositionsError(f"From file does not exist: {from_path}")

	prefixed_files = _prefixed_files(folder_path)
	selected_files = [file_info for file_info in prefixed_files if file_info[0] >= start_position]
	if not selected_files:
		raise ShiftPositionsError(f"No files found at or after position {start_position:04d}")

	selected_source_names = {path.name.casefold() for _, path in selected_files}
	selected_source_positions = {position for position, _ in selected_files}
	existing_positions = {position for position, _ in prefixed_files}
	existing_names = {path.name.casefold() for path in folder_path.iterdir() if path.is_file()}
	plan: list[ShiftPlanItem] = []
	target_names: set[str] = set()
	target_positions: set[int] = set()

	for source_position, source_path in selected_files:
		destination_position = source_position + increment
		if destination_position < 1:
			raise ShiftPositionsError(
				f"Shift would move {source_path.name} before position 0001"
			)
		if destination_position > 9999:
			raise ShiftPositionsError(
				f"Shift would move {source_path.name} beyond 9999"
			)

		destination_name = _replace_position(source_path.name, destination_position)
		destination_name_key = destination_name.casefold()
		if destination_position in target_positions:
			raise ShiftPositionsError(f"Shift would create duplicate target position: {destination_position:04d}")
		if destination_position in existing_positions and destination_position not in selected_source_positions:
			raise ShiftPositionsError(f"Target position already exists outside shifted range: {destination_position:04d}")
		if destination_name_key in target_names:
			raise ShiftPositionsError(f"Shift would create duplicate target name: {destination_name}")
		if destination_name_key in existing_names and destination_name_key not in selected_source_names:
			raise ShiftPositionsError(f"Target already exists outside shifted range: {destination_name}")

		target_names.add(destination_name_key)
		target_positions.add(destination_position)
		plan.append(
			ShiftPlanItem(
				source_path=source_path,
				destination_path=folder_path / destination_name,
				source_position=source_position,
				destination_position=destination_position,
			)
		)

	if increment > 0:
		return tuple(sorted(plan, key=lambda item: item.source_position, reverse=True))

	return tuple(sorted(plan, key=lambda item: item.source_position))


def apply_shift(plan: tuple[ShiftPlanItem, ...]) -> None:
	"""Apply a previously validated shift plan."""
	for item in plan:
		item.source_path.rename(item.destination_path)


def plan_spreadsheet_insert(
	spreadsheet_path: str | Path,
	sheet_name: str,
	start_position: int,
	row_count: int,
) -> SpreadsheetInsertPlan:
	"""Build and validate a spreadsheet-row insertion plan."""
	if row_count < 1:
		raise ShiftPositionsError("Spreadsheet row insertion requires a positive increment")

	path = Path(spreadsheet_path)
	if not path.is_file():
		raise ShiftPositionsError(f"Spreadsheet does not exist: {path}")

	insert_before_row = start_position + 1
	if insert_before_row < 2:
		raise ShiftPositionsError("Spreadsheet insertion row must be after the header row")

	try:
		from openpyxl import load_workbook
	except ImportError as exc:
		raise ShiftPositionsError("openpyxl is required to insert spreadsheet rows. Install requirements.txt.") from exc

	workbook = load_workbook(path, read_only=True)
	try:
		if sheet_name not in workbook.sheetnames:
			raise ShiftPositionsError(f"Spreadsheet does not contain sheet '{sheet_name}'")
		sheet = workbook[sheet_name]
		if insert_before_row > sheet.max_row + 1:
			raise ShiftPositionsError(
				f"Insert row {insert_before_row} is beyond sheet end {sheet.max_row + 1}"
			)
	finally:
		workbook.close()

	return SpreadsheetInsertPlan(
		spreadsheet_path=path,
		sheet_name=sheet_name,
		insert_before_row=insert_before_row,
		row_count=row_count,
		backup_path=_backup_path(path),
	)


def apply_spreadsheet_insert(plan: SpreadsheetInsertPlan) -> None:
	"""Insert blank rows into the spreadsheet after first writing a backup."""
	from openpyxl import load_workbook

	shutil.copy2(plan.spreadsheet_path, plan.backup_path)
	workbook = load_workbook(plan.spreadsheet_path)
	try:
		sheet = workbook[plan.sheet_name]
		sheet.insert_rows(plan.insert_before_row, amount=plan.row_count)
		workbook.save(plan.spreadsheet_path)
	finally:
		workbook.close()


def main(argv: list[str] | None = None) -> int:
	args = _parse_args(argv)
	try:
		config = _load_optional_config(args.config)
		folder = _folder_from_args(args, config)
		plan = plan_shift(folder, args.from_file, args.increment)
		spreadsheet_plan = _spreadsheet_plan_from_args(args, config, plan)
	except (ConfigError, ShiftPositionsError) as exc:
		print(f"ERROR {exc}", file=sys.stderr)
		return 1

	for item in plan:
		action = "rename" if args.apply else "would rename"
		print(f"{action}: {item.source_path.name} -> {item.destination_path.name}")

	if spreadsheet_plan is not None:
		action = "insert" if args.apply else "would insert"
		print(
			f"{action}: {spreadsheet_plan.row_count} spreadsheet row(s) before row "
			f"{spreadsheet_plan.insert_before_row} in {spreadsheet_plan.sheet_name}"
		)
		if args.apply:
			print(f"Spreadsheet backup: {spreadsheet_plan.backup_path}")

	if args.apply:
		apply_shift(plan)
		if spreadsheet_plan is not None:
			apply_spreadsheet_insert(spreadsheet_plan)
		print(f"Renamed {len(plan)} files.")
	else:
		print(f"Dry run only. {len(plan)} files would be renamed. Pass --apply to make changes.")

	return 0


def _parse_args(argv: list[str] | None) -> argparse.Namespace:
	parser = argparse.ArgumentParser(description="Shift 4-digit comic position prefixes from a selected file onward.")
	parser.add_argument("from_file", help="Filename or full path to the first file that should shift.")
	parser.add_argument("--config", help="Organizer config JSON. Provides folder, spreadsheet, and sheet defaults.")
	parser.add_argument("--folder", help="Destination folder. Required when from_file is only a filename unless --config is provided.")
	parser.add_argument("--increment", type=int, default=1, help="Amount to add to each position. Defaults to 1.")
	parser.add_argument("--insert-spreadsheet-rows", action="store_true", help="Also insert blank rows into the reading-order spreadsheet.")
	parser.add_argument("--spreadsheet", help="Spreadsheet path. Defaults to config spreadsheet_path when --config is provided.")
	parser.add_argument("--sheet-name", help="Spreadsheet sheet name. Defaults to config sheet_name, then Issue Release Order.")
	parser.add_argument("--apply", action="store_true", help="Actually rename files. Without this, only prints the plan.")
	return parser.parse_args(argv)


def _load_optional_config(config_path: str | None):
	if config_path is None:
		return None

	return load_config(config_path)


def _folder_from_args(args: argparse.Namespace, config) -> Path:
	if args.folder:
		return Path(args.folder)
	if config is not None:
		return config.destination_folder

	return _folder_from_from_file(args.from_file)


def _spreadsheet_plan_from_args(args: argparse.Namespace, config, plan: tuple[ShiftPlanItem, ...]) -> SpreadsheetInsertPlan | None:
	if not args.insert_spreadsheet_rows:
		return None
	if args.increment < 1:
		raise ShiftPositionsError("Spreadsheet row insertion only supports positive increments")

	spreadsheet_path = args.spreadsheet
	if spreadsheet_path is None and config is not None:
		spreadsheet_path = config.spreadsheet_path
	if spreadsheet_path is None:
		raise ShiftPositionsError("--spreadsheet or --config is required with --insert-spreadsheet-rows")

	sheet_name = args.sheet_name
	if sheet_name is None and config is not None:
		sheet_name = config.sheet_name
	if sheet_name is None:
		sheet_name = "Issue Release Order"

	start_position = min(item.source_position for item in plan)
	return plan_spreadsheet_insert(spreadsheet_path, sheet_name, start_position, args.increment)


def _folder_from_from_file(from_file: str) -> Path:
	from_path = Path(from_file)
	if from_path.parent == Path("."):
		raise ShiftPositionsError("--folder is required when from_file is only a filename")

	return from_path.parent


def _resolve_from_file(folder: Path, from_file: str | Path) -> Path:
	from_path = Path(from_file)
	if from_path.parent == Path("."):
		return folder / from_path.name

	return from_path


def _prefixed_files(folder: Path) -> list[tuple[int, Path]]:
	files: list[tuple[int, Path]] = []
	for path in folder.iterdir():
		if not path.is_file():
			continue

		position = _position_from_name(path.name)
		if position is not None:
			files.append((position, path))

	return sorted(files, key=lambda item: item[0])


def _position_from_name(name: str) -> int | None:
	match = POSITION_PREFIX_RE.match(name)
	if match is None:
		return None

	return int(match.group("position"))


def _replace_position(name: str, position: int) -> str:
	match = POSITION_PREFIX_RE.match(name)
	if match is None:
		raise ShiftPositionsError(f"File does not start with a 4-digit position prefix: {name}")

	return f"{position:04d} - {match.group('rest')}"


def _backup_path(path: Path) -> Path:
	candidate = path.with_name(f"{path.stem}.backup{path.suffix}")
	if not candidate.exists():
		return candidate

	index = 1
	while True:
		candidate = path.with_name(f"{path.stem}.backup-{index}{path.suffix}")
		if not candidate.exists():
			return candidate
		index += 1


if __name__ == "__main__":
	raise SystemExit(main())
